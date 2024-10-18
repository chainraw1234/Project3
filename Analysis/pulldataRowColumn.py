import sys
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import calendar

# ตั้งค่า encoding ให้เป็น utf-8
sys.stdout.reconfigure(encoding='utf-8')

# โหลดไฟล์ workbook และ worksheet
wb = load_workbook('Test/LProfitLoss24-Feb-.xlsx', data_only=True)
ws = wb.active

# สร้าง array (list) เพื่อเก็บข้อมูลในคอลัมน์ C
column_c_data = []

# ตัวแปรเพื่อติดตามจำนวนครั้งที่เจอค่า None
none_count = 0

# เริ่มต้น loop ในคอลัมน์ C โดยเช็คว่าเจอค่า None หรือไม่
for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):  # อ่านข้อมูลจากคอลัมน์ C
    cell_value = row[2]  # คอลัมน์ C คือ index 2
    
    if cell_value is None:
        none_count += 1  # นับจำนวนครั้งที่เจอ None
        if none_count >= 20:  # ถ้าเจอ None ติดต่อกัน 20 ครั้ง
            break  # หยุดการ loop
    else:
        none_count = 0  # รีเซ็ตจำนวนครั้งที่เจอ None ถ้าเจอค่าอื่น
    
    column_c_data.append(cell_value)

# ดูว่า total อยู่ที่บรรทัดไหน
count_row = len(column_c_data)

# สร้าง dictionary เพื่อเก็บคอลัมน์ที่ตรงกับชื่อที่เราต้องการหา
column_mapping = {}

# รายการชื่อคอลัมน์ที่เราต้องการหา
keywords = ['วัตถุดิบ','หมู',' ไก่ ไข่','เนื้อ','วัชระ+สุรพล','ลูกชิ้น',
            'กุ้ง','ผัก','เครื่องดื่ม','ไอติม','ขนมจีบ','ของหวาน','ผลไม้']

keywords_invest = ['ปรับปรุง','ไฟฟ้า','ประปา','Air','ปลวก','สถาบัน','Fur','อุปกรณ์',]

keywords_income = ['รายรับอื่นๆ','ขาย']

keywords_cost = ['วัตถุดิบ','อุปกรณ์ใช้สอย','ค่าเช่า ค่าส่วนกลาง นํ้า','ค่าใช้สอย','PR','ค่าไฟ','ค่าแรง','วัสดุสิ้นเปลือง','Ptime','สวัสดิการ','Manage']

keywords_profit = ['กำไรขั้นต้น (หักวัตถุดิบ)','กำไรขั้นต้นรวมรับอื่นๆ','กำไรขั้นต้น (หักอุปกรณ์ใช้สอย วัสดุสิ้นเปลือง ค่าใช้สอย ค่าไฟ ค่าเช่า)',
                   'กำไรขั้นต้น (หักอุปกรณ์ใช้สอย วัสดุสิ้นเปลือง ค่าใช้สอย ค่าไฟ ค่าเช่า ค่าแรง)','EBITDA (Bht)','EBITDA']

keyword_sell = ['จำนวนวันขาย (วัน)','Total Cost ไม่รวม Manage (บาท)','จำนวนขายขั้นต่ำต่อวัน (บาท)','จำนวนขายขั้นต่ำต่อวัน (คน)']

# วนลูปเพื่อหาชื่อคอลัมน์ในแถวที่ 1
for cell in ws[1]:  # อ่านทุกเซลล์ในแถวที่ 1
    if cell.value in keywords:  # ถ้าชื่อคอลัมน์ตรงกับ keyword
        column_mapping[cell.value] = cell.column_letter  # เก็บตัวอักษรคอลัมน์ไว้ใน dictionary
        

# ดึงข้อมูลจากแต่ละคอลัมน์ตามชื่อที่ต้องการ
price_material = {}
total_price_material = 0

for keyword in keywords:
    column_letter = column_mapping.get(keyword)  # ดึงตัวอักษรคอลัมน์จากชื่อคอลัมน์
    if column_letter:
        column_number = column_index_from_string(column_letter)  # แปลงตัวอักษรคอลัมน์เป็นตัวเลข
        value = ws.cell(row=count_row, column=column_number).value  # ดึงค่าจากแถว count_row
        if value is not None:  # ตรวจสอบค่า None
            price_material[keyword] = value  # เก็บค่าใน dictionary
            total_price_material += value  # บวกเข้ากับ total_price

price_invest = {}
total_invest = 0

# ปรับชื่อคอลัมน์โดยลบช่องว่างรอบๆ คำ
for cell in ws[1]:  # อ่านทุกเซลล์ในแถวที่ 1
    if cell.value and cell.value.strip() in keywords_invest:  # ใช้ .strip() เพื่อกำจัดช่องว่าง
        column_mapping[cell.value.strip()] = cell.column_letter  # เก็บตัวอักษรคอลัมน์ไว้ใน dictionary

for keyword in keywords_invest:
    column_letter = column_mapping.get(keyword)  # ดึงตัวอักษรคอลัมน์จากชื่อคอลัมน์
    if column_letter:
        column_number = column_index_from_string(column_letter)  # แปลงตัวอักษรคอลัมน์เป็นตัวเลข
        value = ws.cell(row=count_row, column=column_number).value  # ดึงค่าจากแถว count_row
        if value is not None:  # ตรวจสอบค่า None
            price_invest[keyword] = value
            total_invest+= value

income = {}
total_income = 0

# ตรวจสอบชื่อคอลัมน์ในแถวที่ 1 ของ Excel โดยลบช่องว่างรอบๆ ชื่อ
for cell in ws[1]:
    column_name = cell.value.strip() if cell.value else None  # ลบช่องว่างรอบๆ
    if column_name in keywords_income:  # ตรวจสอบคอลัมน์ว่าตรงกับคีย์เวิร์ดหรือไม่
        column_mapping[column_name] = cell.column_letter  # เก็บคอลัมน์ที่ตรงใน dictionary

for keyword in keywords_income:
    column_letter = column_mapping.get(keyword)  # ดึงตัวอักษรคอลัมน์จากชื่อคอลัมน์
    if column_letter:
        column_number = column_index_from_string(column_letter)  # แปลงตัวอักษรคอลัมน์เป็นตัวเลข
        value = ws.cell(row=count_row, column=column_number).value  # ดึงค่าจากแถว count_row
        if value is not None:  # ตรวจสอบค่า None
            income[keyword] = value  # เก็บค่าใน dictionary
            total_income += value  # บวกเข้ากับ total_price
            
cost = {}
total_cost = 0

for cell in ws[1]:
    column_name = cell.value.strip() if cell.value else None  # ลบช่องว่างรอบๆ
    if column_name in keywords_cost:  # ตรวจสอบคอลัมน์ว่าตรงกับคีย์เวิร์ดหรือไม่
        column_mapping[column_name] = cell.column_letter  # เก็บคอลัมน์ที่ตรงใน dictionary

for keyword in keywords_cost:
    if keyword == 'วัตถุดิบ':
        cost[keyword] = total_price_material
        total_cost += cost['วัตถุดิบ']
    if keyword == 'ค่าเช่า ค่าส่วนกลาง นํ้า':
        cost[keyword] = price_invest['สถาบัน']
        total_cost += cost['ค่าเช่า ค่าส่วนกลาง นํ้า']
    else:
        if keyword == 'วัตถุดิบ':
            continue
        column_letter = column_mapping.get(keyword)  # ดึงตัวอักษรคอลัมน์จากชื่อคอลัมน์
        if column_letter:
            column_number = column_index_from_string(column_letter)  # แปลงตัวอักษรคอลัมน์เป็นตัวเลข
            value = ws.cell(row=count_row, column=column_number).value  # ดึงค่าจากแถว count_row
            if value is not None:  # ตรวจสอบค่า None
                cost[keyword] = value  # เก็บค่าใน dictionary
                total_cost += value  # บวกเข้ากับ total_price


# ตรวจสอบว่า total_price ไม่เป็น 0 เพื่อหลีกเลี่ยงการหารด้วย 0
if total_price_material > 0:
    percent_price_material = {}

    for keyword in price_material:
        percent_price_material[keyword] = (price_material[keyword] / total_price_material) * 100  # คำนวณเปอร์เซ็นต์
        #print(percent_price_material[keyword]) 

if total_invest > 0 :
    percent_invest = {}

    for keyword in price_invest:
        percent_invest[keyword] = (price_invest[keyword] / total_invest) * 100
        #print(percent_invest[keyword])

if (total_income >0 and total_cost >0):
    percent_total_cost = {}
    percent_sell = {}

    for keyword in cost:
        percent_total_cost[keyword] = (cost[keyword] / total_cost)*100
        
    for keyword in cost:
        if(keyword == 'ค่าไฟ'):
            percent_sell[keyword] = ((cost[keyword]-15000) / total_income)*100
        else:
            percent_sell[keyword] = (cost[keyword] / total_income)*100

    profit = {}

    for keyword in keywords_profit:
        if keyword == 'กำไรขั้นต้น (หักวัตถุดิบ)': profit[keyword] = income['ขาย'] - cost['วัตถุดิบ']
        elif keyword == 'กำไรขั้นต้นรวมรับอื่นๆ' : profit[keyword] = profit['กำไรขั้นต้น (หักวัตถุดิบ)'] + income['รายรับอื่นๆ']
        elif keyword == 'กำไรขั้นต้น (หักอุปกรณ์ใช้สอย วัสดุสิ้นเปลือง ค่าใช้สอย ค่าไฟ ค่าเช่า)':
            profit[keyword] = profit['กำไรขั้นต้นรวมรับอื่นๆ'] - cost['ค่าเช่า ค่าส่วนกลาง นํ้า'] - cost['อุปกรณ์ใช้สอย'] - cost['วัสดุสิ้นเปลือง'] - cost['ค่าใช้สอย'] - cost['PR'] - cost['ค่าไฟ']
        elif keyword == 'กำไรขั้นต้น (หักอุปกรณ์ใช้สอย วัสดุสิ้นเปลือง ค่าใช้สอย ค่าไฟ ค่าเช่า ค่าแรง)' :
            profit[keyword] = profit['กำไรขั้นต้น (หักอุปกรณ์ใช้สอย วัสดุสิ้นเปลือง ค่าใช้สอย ค่าไฟ ค่าเช่า)'] - cost['ค่าแรง'] - cost['Ptime'] - cost['สวัสดิการ']
        elif keyword == 'EBITDA (Bht)':
            profit[keyword] = profit['กำไรขั้นต้น (หักอุปกรณ์ใช้สอย วัสดุสิ้นเปลือง ค่าใช้สอย ค่าไฟ ค่าเช่า ค่าแรง)'] - cost['Manage']
        elif keyword == 'EBITDA':
            profit[keyword] = (profit['EBITDA (Bht)']/total_income)*100

    month = ws['B2'].value
    # ตรวจสอบเดือน
    month_number = month.month
    days_in_month = calendar.monthrange(month.year, month_number)[1]

    sell = {}

    for keyword in keyword_sell:
        if keyword == 'จำนวนวันขาย (วัน)' : sell[keyword] = days_in_month
        elif keyword == 'Total Cost ไม่รวม Manage (บาท)':
            sell[keyword] = total_cost - cost['Manage']
        elif keyword == 'จำนวนขายขั้นต่ำต่อวัน (บาท)':
            sell[keyword] = sell['Total Cost ไม่รวม Manage (บาท)']/sell['จำนวนวันขาย (วัน)']
        elif keyword == 'จำนวนขายขั้นต่ำต่อวัน (คน)': 
            sell[keyword] = sell['จำนวนขายขั้นต่ำต่อวัน (บาท)']/298

print(month)
# print(percent_total_cost)
#print(percent_sell)

# แสดงผลข้อมูลที่เก็บได้
#print(total_price_material)
#print(price_material)
#print(price_invest)
#print(price_invest)
#print(total_income)